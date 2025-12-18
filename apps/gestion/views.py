from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Q, Count 
from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.password_validation import validate_password


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import Persona, Causa, Audiencia, Documento, CausaPersona, LogAuditoria, EstadoCausa, Materia, Tribunal, TipoDocumento, Consentimiento
from .forms import (
    PersonaForm, CausaForm, AudienciaForm,
    DocumentoForm, CausaPersonaForm, ConsentimientoForm
)

from .permissions import (
    permiso_requerido, 
    solo_roles_permitidos,
    tiene_permiso,
    es_admin,
    es_externo,
    obtener_rol_usuario,
    puede_ver_causa,
    puede_editar_causa,
)

# Alias para compatibilidad
usuario_tiene_permiso = tiene_permiso

from datetime import datetime, timedelta, date
from django.utils import timezone
from calendar import monthrange

from .validators import validar_archivo

import calendar

from .cache_utils import (
    get_tribunales_activos,
    get_materias_activas,
    get_estados_activos,
    get_responsables_activos,
    get_tipos_documento_activos,
)

# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def dashboard(request):
    from django.db.models import Count
    from django.utils import timezone
    from datetime import timedelta
    
    hoy = timezone.now().date()
    
    # FILTRO POR ROL: Base queryset según rol
    rol_usuario = obtener_rol_usuario(request.user)
    es_estudiante = rol_usuario == 'ESTUDIANTE'
    
    # Base querysets filtrados para estudiante
    if es_estudiante:
        causas_qs = Causa.objects.filter(responsable=request.user)
        audiencias_qs = Audiencia.objects.filter(causa__responsable=request.user)
        documentos_qs = Documento.objects.filter(causa__responsable=request.user)
    else:
        causas_qs = Causa.objects.all()
        audiencias_qs = Audiencia.objects.all()
        documentos_qs = Documento.objects.all()
    
    # Estadísticas generales
    total_causas = causas_qs.count()
    total_personas = Persona.objects.count()  # Personas no se filtran
    total_audiencias = audiencias_qs.count()
    total_documentos = documentos_qs.count()
    
    # Causas por estado (con objetos completos para el template)
    causas_por_estado_raw = causas_qs.values(
        'estado__nombre', 'estado__color'
    ).annotate(count=Count('id')).order_by('-count')
    
    # Convertir a lista y calcular porcentajes
    causas_por_estado = []
    for item in causas_por_estado_raw:
        # Obtener el objeto estado para el template
        estado_obj = EstadoCausa.objects.filter(nombre=item['estado__nombre']).first()
        if estado_obj:
            porcentaje = round((item['count'] / total_causas * 100) if total_causas > 0 else 0, 1)
            causas_por_estado.append({
                'estado': estado_obj,
                'count': item['count'],
                'porcentaje': porcentaje
            })
    
    # Próximas audiencias
    proximas_audiencias = audiencias_qs.select_related(
        'causa', 'causa__tribunal'
    ).filter(
        fecha_hora__gte=timezone.now(),
        estado__in=['PROGRAMADA', 'CONFIRMADA']
    ).order_by('fecha_hora')[:5]
    
    # Causas recientes
    causas_recientes = causas_qs.select_related(
        'tribunal', 'materia', 'estado', 'responsable'
    ).order_by('-fecha_creacion')[:5]
    
    # Personas recientes (no filtradas - todos ven todas)
    personas_recientes = Persona.objects.order_by('-id')[:5]
    
    # Actividad reciente
    if es_estudiante:
        # Estudiante solo ve actividad de sus causas
        actividad_reciente = LogAuditoria.objects.select_related(
            'usuario'
        ).filter(usuario=request.user).order_by('-fecha')[:10]
    else:
        actividad_reciente = LogAuditoria.objects.select_related(
            'usuario'
        ).order_by('-fecha')[:10]
    
    # Audiencias del mes
    audiencias_mes = audiencias_qs.filter(
        fecha_hora__year=hoy.year,
        fecha_hora__month=hoy.month
    ).count()
    
    # Documentos del mes
    documentos_mes = documentos_qs.filter(
        fecha_subida__year=hoy.year,
        fecha_subida__month=hoy.month
    ).count()
    
    context = {
        'total_causas': total_causas,
        'total_personas': total_personas,
        'total_audiencias': total_audiencias,
        'total_documentos': total_documentos,
        'causas_por_estado': causas_por_estado,
        'proximas_audiencias': proximas_audiencias,
        'causas_recientes': causas_recientes,
        'personas_recientes': personas_recientes,
        'actividad_reciente': actividad_reciente,
        'audiencias_mes': audiencias_mes,
        'documentos_mes': documentos_mes,
    }
    return render(request, 'gestion/dashboard.html', context)


# =============================================================================
# PERSONAS
# =============================================================================

@login_required
def personas_lista(request):
    personas = Persona.objects.all().order_by('-id')
    
    # Filtros
    q = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '')
    
    if q:
        personas = personas.filter(
            Q(nombres__icontains=q) |
            Q(apellidos__icontains=q) |
            Q(run__icontains=q) |
            Q(email__icontains=q)
        )
    
    if tipo:
        personas = personas.filter(tipo_persona=tipo)
    
    # Paginación
    paginator = Paginator(personas, 15)  # 15 por página
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'personas': page_obj,
        'page_obj': page_obj,
        'q': q,
        'tipo': tipo,
    }
    return render(request, 'gestion/personas_lista.html', context)


@permiso_requerido('puede_crear_persona')
@login_required
def persona_crear(request):
    if request.method == 'POST':
        form = PersonaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Persona creada exitosamente.')
            return redirect('gestion:personas_lista')
    else:
        form = PersonaForm()
    
    context = {
        'form': form,
    }
    return render(request, 'gestion/persona_form.html', context)

@permiso_requerido('puede_editar_persona')
def persona_editar(request, pk):
    persona = get_object_or_404(Persona, pk=pk)
    if request.method == 'POST':
        form = PersonaForm(request.POST, instance=persona)
        if form.is_valid():
            form.save()
            messages.success(request, 'Persona actualizada exitosamente.')
            return redirect('gestion:persona_detalle', pk=pk)
    else:
        form = PersonaForm(instance=persona)
    return render(request, 'gestion/persona_form.html', {'form': form, 'persona': persona})

@login_required
def persona_detalle(request, pk):
    persona = get_object_or_404(Persona, pk=pk)
    
    # Causas relacionadas (optimizado)
    causas_relacionadas = CausaPersona.objects.select_related(
        'causa', 'causa__estado', 'causa__tribunal', 'causa__materia'
    ).filter(persona=persona)
    
    # Documentos relacionados (a través de las causas)
    causas_ids = causas_relacionadas.values_list('causa_id', flat=True)
    documentos = Documento.objects.select_related(
        'causa', 'tipo'
    ).filter(causa_id__in=causas_ids).order_by('-fecha_subida')[:5]
    
    # Próximas audiencias de sus causas
    proximas_audiencias = Audiencia.objects.select_related(
        'causa'
    ).filter(
        causa_id__in=causas_ids,
        fecha_hora__gte=timezone.now(),
        estado__in=['PROGRAMADA', 'CONFIRMADA']
    ).order_by('fecha_hora')[:3]
    
    context = {
        'persona': persona,
        'causas_relacionadas': causas_relacionadas,
        'documentos': documentos,
        'proximas_audiencias': proximas_audiencias,
    }
    return render(request, 'gestion/persona_detalle.html', context)


# =============================================================================
# CAUSAS
# =============================================================================

@login_required
def causas_lista(request):
    causas = Causa.objects.select_related('tribunal', 'materia', 'estado', 'responsable').order_by('-fecha_creacion')
    
    # FILTRO POR ROL: Estudiante solo ve causas asignadas a él
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE':
        causas = causas.filter(responsable=request.user)
    
    # Filtros de búsqueda
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '')
    materia = request.GET.get('materia', '')
    
    if q:
        causas = causas.filter(
            Q(caratula__icontains=q) |
            Q(rit__icontains=q) |
            Q(ruc__icontains=q)
        )
    
    if estado:
        causas = causas.filter(estado_id=estado)
    
    if materia:
        causas = causas.filter(materia_id=materia)
    
    # Paginación
    paginator = Paginator(causas, 15)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'causas': page_obj,
        'page_obj': page_obj,
        'estados': EstadoCausa.objects.filter(activo=True),
        'materias': Materia.objects.filter(activo=True),
        'q': q,
        'estado_filtro': estado,
        'materia_filtro': materia,
    }
    return render(request, 'gestion/causas_lista.html', context)

@permiso_requerido('puede_crear_causa')
@login_required
def causa_crear(request):
    rol_usuario = obtener_rol_usuario(request.user)
    
    if request.method == 'POST':
        form = CausaForm(request.POST)
        if form.is_valid():
            causa = form.save(commit=False)
            # Si es estudiante, se autoasigna la causa
            if rol_usuario == 'ESTUDIANTE':
                causa.responsable = request.user
            causa.save()
            messages.success(request, 'Causa creada exitosamente.')
            return redirect('gestion:causas_lista')
    else:
        form = CausaForm()
    
    # Usar caché para catálogos
    context = {
        'form': form,
        'tribunales': Tribunal.objects.filter(activo=True).order_by('nombre'),
        'materias': Materia.objects.filter(activo=True).order_by('nombre'),
        'estados': EstadoCausa.objects.filter(activo=True).order_by('orden'),
        'responsables': User.objects.filter(is_active=True).order_by('first_name', 'username'),
        'es_estudiante': rol_usuario == 'ESTUDIANTE',
        'usuario_actual': request.user,
    }
    return render(request, 'gestion/causa_form.html', context)

@permiso_requerido('puede_editar_causa')
def causa_editar(request, pk):
    causa = get_object_or_404(Causa, pk=pk)
    
    # VALIDACIÓN: Estudiante solo puede editar sus propias causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE' and causa.responsable != request.user:
        messages.error(request, 'Solo puedes editar las causas que tienes asignadas.')
        return redirect('gestion:causas_lista')
    
    if request.method == 'POST':
        form = CausaForm(request.POST, instance=causa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Causa actualizada exitosamente.')
            return redirect('gestion:causa_detalle', pk=pk)
    else:
        form = CausaForm(instance=causa)
    
    # Usar caché para catálogos
    context = {
        'form': form,
        'causa': causa,
        'tribunales': Tribunal.objects.filter(activo=True).order_by('nombre'),
        'materias': Materia.objects.filter(activo=True).order_by('nombre'),
        'estados': EstadoCausa.objects.filter(activo=True).order_by('orden'),
        'responsables': User.objects.filter(is_active=True).order_by('first_name', 'username'),
    }
    return render(request, 'gestion/causa_form.html', context)

@login_required
def causa_detalle(request, pk):
    causa = get_object_or_404(
        Causa.objects.select_related('tribunal', 'materia', 'estado', 'responsable'),
        pk=pk
    )
    
    # VALIDACIÓN: Estudiante solo puede ver sus propias causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE' and causa.responsable != request.user:
        messages.error(request, 'Solo puedes ver las causas que tienes asignadas.')
        return redirect('gestion:causas_lista')
    
    # Personas asociadas (optimizado)
    personas_asociadas = CausaPersona.objects.select_related(
        'persona'
    ).filter(causa=causa)
    
    # Audiencias (optimizado)
    audiencias = Audiencia.objects.filter(causa=causa).order_by('-fecha_hora')
    
    # Documentos (optimizado)
    documentos = Documento.objects.select_related(
        'tipo', 'usuario'
    ).filter(causa=causa).order_by('-fecha_subida')
    
    # Próxima audiencia
    proxima_audiencia = Audiencia.objects.filter(
        causa=causa,
        fecha_hora__gte=timezone.now(),
        estado__in=['PROGRAMADA', 'CONFIRMADA']
    ).order_by('fecha_hora').first()
    
    # Días desde ingreso
    dias_desde_ingreso = (date.today() - causa.fecha_creacion).days
    
    # Actividad reciente (logs de la causa)
    actividad = LogAuditoria.objects.select_related(
        'usuario'
    ).filter(
        modelo='CAUSA',
        objeto_id=causa.pk
    ).order_by('-fecha')[:5]
    
    context = {
        'causa': causa,
        'personas_asociadas': personas_asociadas,
        'audiencias': audiencias,
        'documentos': documentos,
        'proxima_audiencia': proxima_audiencia,
        'dias_desde_ingreso': dias_desde_ingreso,
        'actividad': actividad,
    }
    return render(request, 'gestion/causa_detalle.html', context)

@login_required
@login_required
def causa_persona_crear(request):
    causa_preseleccionada = request.GET.get('causa', '')
    
    # FILTRO POR ROL: Estudiante solo ve sus causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE':
        causas_disponibles = Causa.objects.filter(responsable=request.user).exclude(estado__es_final=True).order_by('-fecha_creacion')
    else:
        causas_disponibles = Causa.objects.exclude(estado__es_final=True).order_by('-fecha_creacion')
    
    if request.method == 'POST':
        form = CausaPersonaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Persona asociada a la causa exitosamente.')
            # Redirigir a la causa si venía de ahí
            causa_id = request.POST.get('causa')
            if causa_id:
                return redirect('gestion:causa_detalle', pk=causa_id)
            return redirect('gestion:causas_lista')
    else:
        form = CausaPersonaForm()
    
    context = {
        'form': form,
        'causas': causas_disponibles,
        'personas': Persona.objects.all().order_by('apellidos', 'nombres'),
        'causa_preseleccionada': causa_preseleccionada,
    }
    return render(request, 'gestion/causa_persona_form.html', context)

@login_required
def causa_persona_editar(request, pk):
    causa_persona = get_object_or_404(CausaPersona, pk=pk)
    
    if request.method == 'POST':
        persona_id = request.POST.get('persona')
        rol_en_causa = request.POST.get('rol_en_causa')
        
        if persona_id and rol_en_causa:
            causa_persona.persona_id = persona_id
            causa_persona.rol_en_causa = rol_en_causa
            causa_persona.save()
            messages.success(request, 'Vinculación actualizada exitosamente.')
            return redirect('gestion:causa_detalle', pk=causa_persona.causa.pk)
    
    context = {
        'causa_persona': causa_persona,
        'personas': Persona.objects.filter(activo=True).order_by('apellidos', 'nombres'),
    }
    return render(request, 'gestion/causa_persona_editar.html', context)


@login_required
def causa_persona_eliminar(request, pk):
    causa_persona = get_object_or_404(CausaPersona, pk=pk)
    causa_id = causa_persona.causa.pk
    causa_persona.delete()
    messages.success(request, 'Persona desvinculada de la causa exitosamente.')
    return redirect('gestion:causa_detalle', pk=causa_id)


# =============================================================================
# AUDIENCIAS
# =============================================================================

@login_required
def audiencias_lista(request):
    audiencias = Audiencia.objects.select_related('causa', 'causa__tribunal', 'causa__responsable').order_by('-fecha_hora')
    
    # FILTRO POR ROL: Estudiante solo ve audiencias de sus causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE':
        audiencias = audiencias.filter(causa__responsable=request.user)
    
    # Filtros de búsqueda
    estado = request.GET.get('estado', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if estado:
        audiencias = audiencias.filter(estado=estado)
    
    if fecha_desde:
        audiencias = audiencias.filter(fecha_hora__date__gte=fecha_desde)
    
    if fecha_hasta:
        audiencias = audiencias.filter(fecha_hora__date__lte=fecha_hasta)
    
    # Paginación
    paginator = Paginator(audiencias, 15)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'audiencias': page_obj,
        'page_obj': page_obj,
        'estado_filtro': estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'gestion/audiencias_lista.html', context)


@login_required
def audiencia_detalle(request, pk):
    audiencia = get_object_or_404(Audiencia, pk=pk)
    
    context = {
        'audiencia': audiencia,
        'puede_editar': usuario_tiene_permiso(request.user, 'puede_editar_audiencia'),
    }
    return render(request, 'gestion/audiencia_detalle.html', context)


@permiso_requerido('puede_crear_audiencia')
def audiencia_crear(request):
    causa_preseleccionada = request.GET.get('causa', '')
    
    if request.method == 'POST':
        form = AudienciaForm(request.POST)
        if form.is_valid():
            audiencia = form.save(commit=False)
            audiencia.creado_por = request.user
            audiencia.save()
            messages.success(request, 'Audiencia creada exitosamente.')
            return redirect('gestion:audiencias_lista')
    else:
        form = AudienciaForm()
    
    # FILTRO POR ROL: Estudiante solo ve sus causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE':
        causas_disponibles = Causa.objects.filter(responsable=request.user).exclude(estado__es_final=True).order_by('-fecha_creacion')
    else:
        causas_disponibles = Causa.objects.exclude(estado__es_final=True).order_by('-fecha_creacion')
    
    context = {
        'form': form,
        'causas': causas_disponibles,
        'causa_preseleccionada': causa_preseleccionada,
    }
    return render(request, 'gestion/audiencia_form.html', context)


@permiso_requerido('puede_editar_audiencia')
def audiencia_editar(request, pk):
    audiencia = get_object_or_404(Audiencia, pk=pk)
    
    # VALIDACIÓN: Estudiante solo puede editar audiencias de sus causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE' and audiencia.causa.responsable != request.user:
        messages.error(request, 'Solo puedes editar audiencias de tus causas asignadas.')
        return redirect('gestion:audiencias_lista')
    
    if request.method == 'POST':
        form = AudienciaForm(request.POST, instance=audiencia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Audiencia actualizada exitosamente.')
            return redirect('gestion:audiencia_detalle', pk=pk)
    else:
        form = AudienciaForm(instance=audiencia)
    
    # FILTRO POR ROL: Estudiante solo ve sus causas
    if rol_usuario == 'ESTUDIANTE':
        causas_disponibles = Causa.objects.filter(responsable=request.user).order_by('-fecha_creacion')
    else:
        causas_disponibles = Causa.objects.all().order_by('-fecha_creacion')
    
    context = {
        'form': form,
        'audiencia': audiencia,
        'causas': causas_disponibles,
    }
    return render(request, 'gestion/audiencia_form.html', context)

# =============================================================================
# DOCUMENTOS
# =============================================================================

@login_required
def documentos_lista(request):
    documentos = Documento.objects.select_related('causa', 'causa__responsable', 'tipo', 'usuario').order_by('-fecha_subida')
    
    # FILTRO POR ROL: Estudiante solo ve documentos de sus causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE':
        documentos = documentos.filter(causa__responsable=request.user)
    
    # Filtros de búsqueda
    q = request.GET.get('q', '').strip()
    tipo = request.GET.get('tipo', '')
    
    if q:
        documentos = documentos.filter(
            Q(titulo__icontains=q) |
            Q(causa__caratula__icontains=q)
        )
    
    if tipo:
        documentos = documentos.filter(tipo_id=tipo)
    
    # Paginación
    paginator = Paginator(documentos, 15)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'documentos': page_obj,
        'page_obj': page_obj,
        'tipos': TipoDocumento.objects.filter(activo=True),
        'q': q,
        'tipo_filtro': tipo,
    }
    return render(request, 'gestion/documentos_lista.html', context)


@login_required
def documento_detalle(request, pk):
    documento = get_object_or_404(Documento, pk=pk)
    
    context = {
        'documento': documento,
    }
    return render(request, 'gestion/documento_detalle.html', context)


@permiso_requerido('puede_subir_documento')
@login_required
def documento_crear(request):
    causa_id = request.GET.get('causa')
    
    if request.method == 'POST':
        causa_id = request.POST.get('causa')
        tipo_id = request.POST.get('tipo')
        titulo = request.POST.get('titulo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        archivo = request.FILES.get('archivo')
        
        # Validaciones
        errores = []
        
        if not causa_id:
            errores.append('Debes seleccionar una causa.')
        
        if not titulo:
            errores.append('El título es obligatorio.')
        
        if not archivo:
            errores.append('Debes seleccionar un archivo.')
        else:
            # Validar archivo
            try:
                validar_archivo(archivo, tipo='documento', max_size=10*1024*1024)
            except ValidationError as e:
                errores.append(str(e.message))
        
        if errores:
            for error in errores:
                messages.error(request, error)
            
            # FILTRO POR ROL: Estudiante solo ve sus causas
            rol_usuario = obtener_rol_usuario(request.user)
            if rol_usuario == 'ESTUDIANTE':
                causas_disponibles = Causa.objects.filter(responsable=request.user).order_by('-fecha_creacion')
            else:
                causas_disponibles = Causa.objects.all().order_by('-fecha_creacion')
            
            return render(request, 'gestion/documento_form.html', {
                'causas': causas_disponibles,
                'tipos_documento': TipoDocumento.objects.filter(activo=True),
                'causa_preseleccionada': causa_id,
            })
        
        # Crear documento
        documento = Documento(
            causa_id=causa_id,
            tipo_id=tipo_id if tipo_id else None,
            titulo=titulo,
            descripcion=descripcion,
            archivo=archivo,
            usuario=request.user,
            estado=request.POST.get('estado', 'PENDIENTE'),
            es_confidencial=request.POST.get('es_confidencial') == 'on',
            folio=request.POST.get('folio', ''),
            fecha_emision=request.POST.get('fecha_emision') or None,
            numero_documento=request.POST.get('numero_documento', ''),
            emisor=request.POST.get('emisor', ''),
        )
        documento.save()
        
        messages.success(request, 'Documento subido exitosamente.')
        
        # Redirigir a la causa si venía de ahí
        if causa_id:
            return redirect('gestion:causa_detalle', pk=causa_id)
        return redirect('gestion:documentos_lista')
    
    # FILTRO POR ROL: Estudiante solo ve sus causas
    rol_usuario = obtener_rol_usuario(request.user)
    if rol_usuario == 'ESTUDIANTE':
        causas_disponibles = Causa.objects.filter(responsable=request.user).order_by('-fecha_creacion')
    else:
        causas_disponibles = Causa.objects.all().order_by('-fecha_creacion')
    
    context = {
        'causas': causas_disponibles,
        'tipos_documento': TipoDocumento.objects.filter(activo=True),
        'causa_preseleccionada': causa_id,
    }
    return render(request, 'gestion/documento_form.html', context)

# =============================================================================
# RELACIÓN CAUSA - PERSONA
# =============================================================================

# =============================================================================
# BÚSQUEDA
# =============================================================================

@login_required
def buscar(request):
    query = request.GET.get('q', '').strip()
    en_causas = request.GET.get('en_causas')
    en_personas = request.GET.get('en_personas')
    en_documentos = request.GET.get('en_documentos')
    estado_filtro = request.GET.get('estado', '')
    materia_filtro = request.GET.get('materia', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Si no hay ningún checkbox marcado y hay query, buscar en todos
    if query and not en_causas and not en_personas and not en_documentos:
        en_causas = True
        en_personas = True
        en_documentos = True
    
    causas_encontradas = []
    personas_encontradas = []
    documentos_encontrados = []
    
    # FILTRO POR ROL: Estudiante solo ve sus datos
    rol_usuario = obtener_rol_usuario(request.user)
    es_estudiante = rol_usuario == 'ESTUDIANTE'
    
    if query:
        # Búsqueda en causas
        if en_causas:
            causas = Causa.objects.select_related('tribunal', 'materia', 'estado', 'responsable')
            
            # Filtro estudiante
            if es_estudiante:
                causas = causas.filter(responsable=request.user)
            
            causas = causas.filter(
                Q(rit__icontains=query) |
                Q(ruc__icontains=query) |
                Q(caratula__icontains=query)
            )
            
            if estado_filtro:
                causas = causas.filter(estado_id=estado_filtro)
            if materia_filtro:
                causas = causas.filter(materia_id=materia_filtro)
            if fecha_desde:
                causas = causas.filter(fecha_creacion__gte=fecha_desde)
            if fecha_hasta:
                causas = causas.filter(fecha_creacion__lte=fecha_hasta)
            
            causas_encontradas = list(causas[:50])
        
        # Búsqueda en personas
        if en_personas:
            personas = Persona.objects.filter(
                Q(run__icontains=query) |
                Q(nombres__icontains=query) |
                Q(apellidos__icontains=query) |
                Q(email__icontains=query)
            )
            personas_encontradas = list(personas[:50])
        
        # Búsqueda en documentos
        if en_documentos:
            documentos = Documento.objects.select_related('causa', 'causa__responsable', 'tipo')
            
            # Filtro estudiante
            if es_estudiante:
                documentos = documentos.filter(causa__responsable=request.user)
            
            documentos = documentos.filter(
                Q(titulo__icontains=query) |
                Q(descripcion__icontains=query) |
                Q(numero_documento__icontains=query)
            )
            documentos_encontrados = list(documentos[:50])
    
    total_resultados = len(causas_encontradas) + len(personas_encontradas) + len(documentos_encontrados)
    
    context = {
        'query': query,
        'en_causas': en_causas,
        'en_personas': en_personas,
        'en_documentos': en_documentos,
        'estado_filtro': estado_filtro,
        'materia_filtro': materia_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'causas_encontradas': causas_encontradas,
        'personas_encontradas': personas_encontradas,
        'documentos_encontrados': documentos_encontrados,
        'total_resultados': total_resultados,
        'estados': EstadoCausa.objects.filter(activo=True).order_by('orden'),
        'materias': Materia.objects.filter(activo=True).order_by('nombre'),
    }
    return render(request, 'gestion/buscar.html', context)

@permiso_requerido('puede_ver_auditoria')
@login_required
def auditoria_lista(request):
    # Verificar permisos
    if not request.user.is_superuser:
        if hasattr(request.user, 'perfil'):
            if request.user.perfil.rol not in ['ADMIN', 'DIRECTOR', 'SUPERVISOR']:
                messages.error(request, 'No tienes permisos para ver la auditoría.')
                return redirect('gestion:dashboard')
    
    logs = LogAuditoria.objects.select_related('usuario').order_by('-fecha')
    
    # Filtros
    usuario = request.GET.get('usuario', '')
    accion = request.GET.get('accion', '')
    modelo = request.GET.get('modelo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    if usuario:
        logs = logs.filter(usuario_id=usuario)
    
    if accion:
        logs = logs.filter(accion=accion)
    
    if modelo:
        logs = logs.filter(modelo=modelo)
    
    if fecha_desde:
        logs = logs.filter(fecha__date__gte=fecha_desde)
    
    if fecha_hasta:
        logs = logs.filter(fecha__date__lte=fecha_hasta)
    
    # Estadísticas
    total = logs.count()
    creaciones = logs.filter(accion='CREAR').count()
    ediciones = logs.filter(accion='EDITAR').count()
    eliminaciones = logs.filter(accion='ELIMINAR').count()
    
    # Paginación
    paginator = Paginator(logs, 25)  # 25 por página para auditoría
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    from django.contrib.auth.models import User
    
    context = {
        'logs': page_obj,
        'page_obj': page_obj,
        'usuarios': User.objects.filter(is_active=True).order_by('username'),
        'total': total,
        'creaciones': creaciones,
        'ediciones': ediciones,
        'eliminaciones': eliminaciones,
        'usuario_filtro': usuario,
        'accion_filtro': accion,
        'modelo_filtro': modelo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'gestion/auditoria_lista.html', context)


@permiso_requerido('puede_ver_auditoria')
@login_required
def auditoria_detalle(request, pk):
    log = get_object_or_404(LogAuditoria, pk=pk)
    
    context = {
        'log': log,
    }
    return render(request, 'gestion/auditoria_detalle.html', context)

@login_required
def causa_linea_tiempo(request, pk):
    causa = get_object_or_404(Causa, pk=pk)
    
    eventos = []
    
    # Función auxiliar para convertir date a datetime
    def to_datetime(fecha):
        if isinstance(fecha, datetime):
            return fecha
        # Si es date, convertir a datetime al inicio del día
        return timezone.make_aware(datetime.combine(fecha, datetime.min.time()))
    
    # Evento: Creación de la causa
    eventos.append({
        'fecha': to_datetime(causa.fecha_creacion),
        'tipo': 'CREACION',
        'titulo': 'Causa creada',
        'descripcion': f'Se registró la causa {causa.caratula}',
        'icono': '📁'
    })
    
    # Eventos: Documentos
    for doc in causa.documentos.all():
        eventos.append({
            'fecha': to_datetime(doc.fecha_subida),
            'tipo': 'DOCUMENTO',
            'titulo': f'Documento: {doc.titulo}',
            'descripcion': f'Tipo: {doc.tipo}' + (f' - Folio: {doc.folio}' if doc.folio else ''),
            'icono': '📄',
            'objeto': doc
        })
    
    # Eventos: Audiencias
    for aud in causa.audiencias.all():
        estado_texto = f' ({aud.get_estado_display()})' if aud.estado != 'PROGRAMADA' else ''
        eventos.append({
            'fecha': to_datetime(aud.fecha_hora),
            'tipo': 'AUDIENCIA',
            'titulo': f'{aud.get_tipo_evento_display()}{estado_texto}',
            'descripcion': f'Lugar: {aud.lugar or "Por definir"}',
            'icono': '📅',
            'objeto': aud
        })
    
    # Eventos: Logs de auditoría relacionados con la causa
    logs = LogAuditoria.objects.filter(
        modelo='CAUSA',
        objeto_id=causa.pk,
        accion__in=['EDITAR', 'CAMBIO_ESTADO', 'ASIGNAR']
    )
    for log in logs:
        eventos.append({
            'fecha': to_datetime(log.fecha),
            'tipo': 'CAMBIO',
            'titulo': log.get_accion_display(),
            'descripcion': log.descripcion or 'Modificación en la causa',
            'icono': '✏️'
        })
    
    # Ordenar por fecha descendente
    eventos.sort(key=lambda x: x['fecha'], reverse=True)
    
    # Contadores para el resumen
    audiencias_count = causa.audiencias.count()
    documentos_count = causa.documentos.count()
    
    context = {
        'causa': causa,
        'eventos': eventos,
        'audiencias_count': audiencias_count,
        'documentos_count': documentos_count,
        'puede_editar': usuario_tiene_permiso(request.user, 'puede_editar_causa'),
    }
    return render(request, 'gestion/causa_linea_tiempo.html', context)

# =============================================================================
# CALENDARIO
# =============================================================================


@login_required
def calendario(request):
    import calendar
    from calendar import monthrange
    from django.utils import timezone
    
    # Obtener mes a mostrar
    mes_param = request.GET.get('mes', '')
    if mes_param:
        try:
            year, month = map(int, mes_param.split('-'))
            fecha_actual = date(year, month, 1)
        except:
            fecha_actual = date.today().replace(day=1)
    else:
        fecha_actual = date.today().replace(day=1)
    
    # Calcular mes anterior y siguiente
    if fecha_actual.month == 1:
        mes_anterior = date(fecha_actual.year - 1, 12, 1)
    else:
        mes_anterior = date(fecha_actual.year, fecha_actual.month - 1, 1)
    
    if fecha_actual.month == 12:
        mes_siguiente = date(fecha_actual.year + 1, 1, 1)
    else:
        mes_siguiente = date(fecha_actual.year, fecha_actual.month + 1, 1)
    
    # Audiencias del mes
    audiencias_mes = Audiencia.objects.select_related(
        'causa', 'causa__tribunal'
    ).filter(
        fecha_hora__year=fecha_actual.year,
        fecha_hora__month=fecha_actual.month
    ).order_by('fecha_hora')
    
    # Organizar audiencias por día
    audiencias_por_dia = {}
    for aud in audiencias_mes:
        dia = aud.fecha_hora.day
        if dia not in audiencias_por_dia:
            audiencias_por_dia[dia] = []
        audiencias_por_dia[dia].append({
            'id': aud.id,
            'hora': aud.fecha_hora.strftime('%H:%M'),
            'tipo': aud.get_tipo_evento_display(),
            'causa': str(aud.causa.caratula)[:30],
            'estado': aud.estado,
        })
    
    # Crear calendario
    cal = calendar.Calendar(firstweekday=0)
    semanas = cal.monthdatescalendar(fecha_actual.year, fecha_actual.month)
    
    # Próximas audiencias
    proximas = Audiencia.objects.select_related(
        'causa'
    ).filter(
        fecha_hora__gte=timezone.now(),
        estado__in=['PROGRAMADA', 'CONFIRMADA']
    ).order_by('fecha_hora')[:10]
    
    # Estadísticas del mes
    total_mes = audiencias_mes.count()
    programadas_mes = audiencias_mes.filter(estado='PROGRAMADA').count()
    confirmadas_mes = audiencias_mes.filter(estado='CONFIRMADA').count()
    realizadas_mes = audiencias_mes.filter(estado='REALIZADA').count()
    
    context = {
        'fecha_actual': fecha_actual,
        'mes_anterior': mes_anterior,
        'mes_siguiente': mes_siguiente,
        'semanas': semanas,
        'audiencias_por_dia': audiencias_por_dia,
        'proximas': proximas,
        'total_mes': total_mes,
        'programadas_mes': programadas_mes,
        'confirmadas_mes': confirmadas_mes,
        'realizadas_mes': realizadas_mes,
        'hoy': date.today(),
    }
    return render(request, 'gestion/calendario.html', context)

# =============================================================================
# REPORTES
# =============================================================================

@permiso_requerido('puede_ver_reportes')
@login_required
def reportes(request):
    # Filtros
    estado = request.GET.get('estado', '')
    materia = request.GET.get('materia', '')
    tribunal = request.GET.get('tribunal', '')
    responsable = request.GET.get('responsable', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Query base optimizada
    causas = Causa.objects.select_related(
        'tribunal', 'materia', 'estado', 'responsable'
    )
    
    # Aplicar filtros
    if estado:
        causas = causas.filter(estado_id=estado)
    if materia:
        causas = causas.filter(materia_id=materia)
    if tribunal:
        causas = causas.filter(tribunal_id=tribunal)
    if responsable:
        causas = causas.filter(responsable_id=responsable)
    if fecha_desde:
        causas = causas.filter(fecha_creacion__gte=fecha_desde)
    if fecha_hasta:
        causas = causas.filter(fecha_creacion__lte=fecha_hasta)
    
    # Estadísticas
    total = causas.count()
    
    # Causas por estado (para gráfico)
    causas_por_estado_qs = causas.values(
        'estado__nombre', 'estado__color'
    ).annotate(total=Count('id')).order_by('-total')
    
    # Convertir a lista y calcular porcentajes
    causas_por_estado = []
    for item in causas_por_estado_qs:
        porcentaje = round((item['total'] / total * 100) if total > 0 else 0, 1)
        causas_por_estado.append({
            'estado__nombre': item['estado__nombre'],
            'estado__color': item['estado__color'],
            'total': item['total'],
            'porcentaje': porcentaje
        })
    
    # Causas por materia (para gráfico)
    causas_por_materia_qs = causas.values(
        'materia__nombre'
    ).annotate(total=Count('id')).order_by('-total')[:5]
    
    # Convertir a lista y calcular porcentajes
    causas_por_materia = []
    for item in causas_por_materia_qs:
        porcentaje = round((item['total'] / total * 100) if total > 0 else 0, 1)
        causas_por_materia.append({
            'materia__nombre': item['materia__nombre'],
            'total': item['total'],
            'porcentaje': porcentaje
        })
    
    # Causas en tramitación vs finalizadas
    en_tramitacion = causas.filter(estado__es_final=False).count()
    finalizadas = causas.filter(estado__es_final=True).count()
    tasa_exito = round((finalizadas / total * 100) if total > 0 else 0, 1)
    
    # Datos para la tabla (limitado a 100)
    causas_tabla = causas.order_by('-fecha_creacion')[:100]
    
    from django.contrib.auth.models import User
    
    context = {
        'causas': causas_tabla,
        'total': total,
        'en_tramitacion': en_tramitacion,
        'finalizadas': finalizadas,
        'tasa_exito': tasa_exito,
        'causas_por_estado': causas_por_estado,
        'causas_por_materia': causas_por_materia,
        'estados': EstadoCausa.objects.filter(activo=True),
        'materias': Materia.objects.filter(activo=True),
        'tribunales': Tribunal.objects.filter(activo=True),
        'responsables': User.objects.filter(is_active=True),
        'estado_filtro': estado,
        'materia_filtro': materia,
        'tribunal_filtro': tribunal,
        'responsable_filtro': responsable,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    return render(request, 'gestion/reportes.html', context)

@solo_roles_permitidos('ADMIN', 'DIRECTOR', 'SUPERVISOR')
def exportar_causas_excel(request):
    """Exporta causas a Excel con los filtros aplicados."""
    # Aplicar mismos filtros que en la vista de reportes
    estado_id = request.GET.get('estado', '')
    materia_id = request.GET.get('materia', '')
    tribunal_id = request.GET.get('tribunal', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    responsable_id = request.GET.get('responsable', '')
    
    causas = Causa.objects.select_related(
        'tribunal', 'materia', 'estado', 'responsable'
    ).all()
    
    if estado_id:
        causas = causas.filter(estado_id=estado_id)
    if materia_id:
        causas = causas.filter(materia_id=materia_id)
    if tribunal_id:
        causas = causas.filter(tribunal_id=tribunal_id)
    if fecha_desde:
        causas = causas.filter(fecha_creacion__gte=fecha_desde)
    if fecha_hasta:
        causas = causas.filter(fecha_creacion__lte=fecha_hasta)
    if responsable_id:
        causas = causas.filter(responsable_id=responsable_id)
    
    causas = causas.order_by('-fecha_creacion')
    
    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Causas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = [
        'RIT', 'RUC', 'Carátula', 'Tribunal', 'Materia', 
        'Estado', 'Responsable', 'Fecha Creación'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row, causa in enumerate(causas, 2):
        ws.cell(row=row, column=1, value=causa.rit or '-').border = thin_border
        ws.cell(row=row, column=2, value=causa.ruc or '-').border = thin_border
        ws.cell(row=row, column=3, value=causa.caratula).border = thin_border
        ws.cell(row=row, column=4, value=str(causa.tribunal) if causa.tribunal else '-').border = thin_border
        ws.cell(row=row, column=5, value=str(causa.materia) if causa.materia else '-').border = thin_border
        ws.cell(row=row, column=6, value=causa.estado.nombre if causa.estado else '-').border = thin_border
        ws.cell(row=row, column=7, value=causa.responsable.get_full_name() if causa.responsable else '-').border = thin_border
        ws.cell(row=row, column=8, value=causa.fecha_creacion.strftime('%d/%m/%Y') if causa.fecha_creacion else '-').border = thin_border
    
    # Ajustar anchos de columna
    column_widths = [15, 20, 40, 30, 25, 15, 25, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=causas_{date.today().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response

@solo_roles_permitidos('ADMIN', 'DIRECTOR', 'SUPERVISOR')
def exportar_causas_pdf(request):
    """Exporta causas a PDF con los filtros aplicados."""
    # Aplicar mismos filtros
    estado_id = request.GET.get('estado', '')
    materia_id = request.GET.get('materia', '')
    tribunal_id = request.GET.get('tribunal', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    responsable_id = request.GET.get('responsable', '')
    
    causas = Causa.objects.select_related(
        'tribunal', 'materia', 'estado', 'responsable'
    ).all()
    
    if estado_id:
        causas = causas.filter(estado_id=estado_id)
    if materia_id:
        causas = causas.filter(materia_id=materia_id)
    if tribunal_id:
        causas = causas.filter(tribunal_id=tribunal_id)
    if fecha_desde:
        causas = causas.filter(fecha_creacion__gte=fecha_desde)
    if fecha_hasta:
        causas = causas.filter(fecha_creacion__lte=fecha_hasta)
    if responsable_id:
        causas = causas.filter(responsable_id=responsable_id)
    
    causas = causas.order_by('-fecha_creacion')[:100]  # Limitar a 100 para PDF
    
    # Crear respuesta PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=causas_{date.today().strftime("%Y%m%d")}.pdf'
    
    # Crear documento
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1  # Centrado
    )
    elements.append(Paragraph("Reporte de Causas - Clínica Jurídica USS", title_style))
    
    # Fecha de generación
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.gray,
        alignment=1,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Generado el {date.today().strftime('%d/%m/%Y')}", date_style))
    elements.append(Spacer(1, 20))
    
    # Datos de la tabla
    data = [['RIT', 'Carátula', 'Tribunal', 'Materia', 'Estado', 'Responsable', 'Fecha']]
    
    for causa in causas:
        data.append([
            causa.rit or '-',
            (causa.caratula[:30] + '...') if len(causa.caratula) > 30 else causa.caratula,
            (str(causa.tribunal)[:20] + '...') if causa.tribunal and len(str(causa.tribunal)) > 20 else str(causa.tribunal or '-'),
            (str(causa.materia)[:15] + '...') if causa.materia and len(str(causa.materia)) > 15 else str(causa.materia or '-'),
            causa.estado.nombre if causa.estado else '-',
            (causa.responsable.get_full_name()[:15] + '...') if causa.responsable and len(causa.responsable.get_full_name()) > 15 else (causa.responsable.get_full_name() if causa.responsable else '-'),
            causa.fecha_creacion.strftime('%d/%m/%Y') if causa.fecha_creacion else '-'
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[60, 120, 100, 80, 70, 90, 60])
    
    # Estilos de tabla
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        # Cuerpo
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        
        # Filas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]))
    
    elements.append(table)
    
    # Pie de página
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray,
        alignment=1
    )
    elements.append(Paragraph(f"Total: {len(data) - 1} causas", footer_style))
    
    # Construir PDF
    doc.build(elements)
    
    return response

@login_required
def perfil(request):
    # Estadísticas del usuario
    mis_causas = Causa.objects.filter(responsable=request.user).count()
    mis_audiencias = Audiencia.objects.filter(
        causa__responsable=request.user,
        fecha_hora__gte=timezone.now(),
        estado__in=['PROGRAMADA', 'CONFIRMADA']
    ).count()
    mis_documentos = Documento.objects.filter(usuario=request.user).count()
    
    # Actividad reciente
    actividad_reciente = LogAuditoria.objects.filter(
        usuario=request.user
    ).order_by('-fecha')[:5]
    
    context = {
        'mis_causas': mis_causas,
        'mis_audiencias': mis_audiencias,
        'mis_documentos': mis_documentos,
        'actividad_reciente': actividad_reciente,
    }
    return render(request, 'gestion/perfil.html', context)


@login_required
def perfil_editar(request):
    if request.method == 'POST':
        # Actualizar datos del usuario
        request.user.email = request.POST.get('email', '')
        request.user.first_name = request.POST.get('first_name', '')
        request.user.last_name = request.POST.get('last_name', '')
        request.user.save()
        
        # Actualizar datos del perfil
        perfil = request.user.perfil
        perfil.rut = request.POST.get('rut', '')
        perfil.telefono = request.POST.get('telefono', '')
        perfil.direccion = request.POST.get('direccion', '')
        perfil.save()
        
        messages.success(request, 'Perfil actualizado exitosamente.')
        return redirect('gestion:perfil')
    
    return render(request, 'gestion/perfil_editar.html')


@login_required
def perfil_cambiar_password(request):
    if request.method == 'POST':
        password_actual = request.POST.get('password_actual', '')
        password_nueva = request.POST.get('password_nueva', '')
        password_confirmar = request.POST.get('password_confirmar', '')
        
        # Verificar contraseña actual
        if not request.user.check_password(password_actual):
            messages.error(request, 'La contraseña actual es incorrecta.')
            return render(request, 'gestion/perfil_cambiar_password.html')
        
        # Verificar que las nuevas coincidan
        if password_nueva != password_confirmar:
            messages.error(request, 'Las contraseñas nuevas no coinciden.')
            return render(request, 'gestion/perfil_cambiar_password.html')
        
        # Verificar requisitos mínimos
        if len(password_nueva) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return render(request, 'gestion/perfil_cambiar_password.html')
        
        # Cambiar contraseña
        request.user.set_password(password_nueva)
        request.user.save()
        
        # Actualizar sesión para no cerrar la sesión
        from django.contrib.auth import update_session_auth_hash
        update_session_auth_hash(request, request.user)
        
        messages.success(request, 'Contraseña actualizada exitosamente.')
        return redirect('gestion:perfil')
    
    return render(request, 'gestion/perfil_cambiar_password.html')

# ============================================================================
# PANEL DE ADMINISTRACIÓN
# ============================================================================

@permiso_requerido('puede_ver_panel_admin')
@login_required
def admin_panel(request):
    context = {
        'total_usuarios': User.objects.count(),
        'total_causas': Causa.objects.count(),
        'total_personas': Persona.objects.count(),
        'total_logs': LogAuditoria.objects.count(),
        'total_estados': EstadoCausa.objects.count(),
        'total_materias': Materia.objects.count(),
        'total_tribunales': Tribunal.objects.count(),
        'total_tipos_documento': TipoDocumento.objects.count(),
        'usuarios_recientes': User.objects.select_related('perfil').order_by('-date_joined')[:5],
    }
    return render(request, 'gestion/admin/panel.html', context)


@permiso_requerido('puede_gestionar_usuarios')
@login_required
def admin_usuarios(request):
    
    # Optimizado con select_related
    usuarios = User.objects.select_related('perfil').order_by('-date_joined')
    
    # Filtros
    filtro = request.GET.get('filtro', '')
    q = request.GET.get('q', '').strip()
    
    if filtro == 'activos':
        usuarios = usuarios.filter(is_active=True)
    elif filtro == 'inactivos':
        usuarios = usuarios.filter(is_active=False)
    elif filtro == 'admin':
        usuarios = usuarios.filter(
            Q(is_superuser=True) | Q(perfil__rol='ADMIN')
        )
    
    if q:
        usuarios = usuarios.filter(
            Q(username__icontains=q) |
            Q(first_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(email__icontains=q)
        )
    
    # Paginación
    paginator = Paginator(usuarios, 15)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'usuarios': page_obj,
        'page_obj': page_obj,
        'filtro': filtro,
        'q': q,
    }
    return render(request, 'gestion/admin/usuarios_lista.html', context)


@permiso_requerido('puede_gestionar_usuarios')
@login_required
def admin_usuario_crear(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        
        # Validaciones
        if User.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya existe.')
            return render(request, 'gestion/admin/usuario_form.html')
        
        if password != password_confirm:
            messages.error(request, 'Las contraseñas no coinciden.')
            return render(request, 'gestion/admin/usuario_form.html')
        
        if len(password) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return render(request, 'gestion/admin/usuario_form.html')
        
        # Crear usuario
        usuario = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=request.POST.get('first_name', ''),
            last_name=request.POST.get('last_name', ''),
            is_active=request.POST.get('is_active') == 'on',
            is_staff=request.POST.get('is_staff') == 'on',
            is_superuser=request.POST.get('is_superuser') == 'on',
        )
        
        # Actualizar perfil
        perfil = usuario.perfil
        perfil.rol = request.POST.get('rol', 'ALUMNO')
        perfil.sede = request.POST.get('sede', '')
        perfil.rut = request.POST.get('rut', '')
        perfil.telefono = request.POST.get('telefono', '')
        perfil.direccion = request.POST.get('direccion', '')
        perfil.save()
        
        messages.success(request, f'Usuario {username} creado exitosamente.')
        return redirect('gestion:admin_usuarios')
    
    return render(request, 'gestion/admin/usuario_form.html')


@permiso_requerido('puede_gestionar_usuarios')
@login_required
def admin_usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    perfil = usuario.perfil if hasattr(usuario, 'perfil') else None
    
    if request.method == 'POST':
        # Actualizar usuario
        usuario.email = request.POST.get('email', '')
        usuario.first_name = request.POST.get('first_name', '')
        usuario.last_name = request.POST.get('last_name', '')
        usuario.is_active = request.POST.get('is_active') == 'on'
        
        if request.user.is_superuser:
            usuario.is_staff = request.POST.get('is_staff') == 'on'
            usuario.is_superuser = request.POST.get('is_superuser') == 'on'
        
        usuario.save()
        
        # Actualizar perfil
        if perfil:
            perfil.rol = request.POST.get('rol', perfil.rol)
            perfil.sede = request.POST.get('sede', '')
            perfil.rut = request.POST.get('rut', '')
            perfil.telefono = request.POST.get('telefono', '')
            perfil.direccion = request.POST.get('direccion', '')
            perfil.save()
        
        messages.success(request, f'Usuario {usuario.username} actualizado exitosamente.')
        return redirect('gestion:admin_usuarios')
    
    context = {
        'usuario': usuario,
        'perfil': perfil,
    }
    return render(request, 'gestion/admin/usuario_form.html', context)


@login_required
def admin_usuario_toggle(request, pk):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos para esta acción.')
        return redirect('gestion:dashboard')
    
    usuario = get_object_or_404(User, pk=pk)
    
    if usuario.pk == request.user.pk:
        messages.error(request, 'No puedes desactivar tu propia cuenta.')
        return redirect('gestion:admin_usuarios')
    
    usuario.is_active = not usuario.is_active
    usuario.save()
    
    estado = 'activado' if usuario.is_active else 'desactivado'
    messages.success(request, f'Usuario {usuario.username} {estado}.')
    return redirect('gestion:admin_usuarios')


# ============================================================================
# CATÁLOGOS
# ============================================================================

CATALOGOS = {
    'estados': {
        'modelo': EstadoCausa,
        'nombre': 'Estados de Causa',
    },
    'materias': {
        'modelo': Materia,
        'nombre': 'Materias',
    },
    'tribunales': {
        'modelo': Tribunal,
        'nombre': 'Tribunales',
    },
    'tipos_documento': {
        'modelo': TipoDocumento,
        'nombre': 'Tipos de Documento',
    },
}


@permiso_requerido('puede_gestionar_catalogos')
@login_required
def admin_catalogo(request, tipo):
    if tipo not in CATALOGOS:
        messages.error(request, 'Catálogo no encontrado.')
        return redirect('gestion:admin_panel')
    
    catalogo = CATALOGOS[tipo]
    items = catalogo['modelo'].objects.all().order_by('nombre')
    
    context = {
        'catalogo_tipo': tipo,
        'catalogo_nombre': catalogo['nombre'],
        'items': items,
    }
    return render(request, 'gestion/admin/catalogo_lista.html', context)


@permiso_requerido('puede_gestionar_catalogos')
@login_required
def admin_catalogo_crear(request, tipo):
    if tipo not in CATALOGOS:
        messages.error(request, 'Catálogo no encontrado.')
        return redirect('gestion:admin_panel')
    
    catalogo = CATALOGOS[tipo]
    
    if request.method == 'POST':
        item = catalogo['modelo']()
        item.nombre = request.POST.get('nombre', '')
        item.activo = request.POST.get('activo') == 'on'
        
        # Campos específicos por tipo
        if tipo == 'estados':
            item.color = request.POST.get('color', 'gray')
            item.orden = int(request.POST.get('orden', 0))
            item.es_final = request.POST.get('es_final') == 'on'
        elif tipo == 'tribunales':
            item.direccion = request.POST.get('direccion', '')
            item.telefono = request.POST.get('telefono', '')
            item.email = request.POST.get('email', '')
        elif tipo == 'tipos_documento':
            item.categoria = request.POST.get('categoria', '')
        elif tipo == 'materias':
            item.descripcion = request.POST.get('descripcion', '')
        
        item.save()
        messages.success(request, f'{catalogo["nombre"]}: registro creado exitosamente.')
        return redirect('gestion:admin_catalogo', tipo=tipo)
    
    context = {
        'catalogo_tipo': tipo,
        'catalogo_nombre': catalogo['nombre'],
    }
    return render(request, 'gestion/admin/catalogo_form.html', context)


@permiso_requerido('puede_gestionar_catalogos')
@login_required
def admin_catalogo_editar(request, tipo, pk):
    if tipo not in CATALOGOS:
        messages.error(request, 'Catálogo no encontrado.')
        return redirect('gestion:admin_panel')
    
    catalogo = CATALOGOS[tipo]
    item = get_object_or_404(catalogo['modelo'], pk=pk)
    
    if request.method == 'POST':
        item.nombre = request.POST.get('nombre', '')
        item.activo = request.POST.get('activo') == 'on'
        
        # Campos específicos por tipo
        if tipo == 'estados':
            item.color = request.POST.get('color', 'gray')
            item.orden = int(request.POST.get('orden', 0))
            item.es_final = request.POST.get('es_final') == 'on'
        elif tipo == 'tribunales':
            item.direccion = request.POST.get('direccion', '')
            item.telefono = request.POST.get('telefono', '')
            item.email = request.POST.get('email', '')
        elif tipo == 'tipos_documento':
            item.categoria = request.POST.get('categoria', '')
        elif tipo == 'materias':
            item.descripcion = request.POST.get('descripcion', '')
        
        item.save()
        messages.success(request, f'{catalogo["nombre"]}: registro actualizado exitosamente.')
        return redirect('gestion:admin_catalogo', tipo=tipo)
    
    context = {
        'catalogo_tipo': tipo,
        'catalogo_nombre': catalogo['nombre'],
        'item': item,
    }
    return render(request, 'gestion/admin/catalogo_form.html', context)


@permiso_requerido('puede_gestionar_catalogos')
@login_required
def admin_catalogo_toggle(request, tipo, pk):
    if tipo not in CATALOGOS:
        messages.error(request, 'Catálogo no encontrado.')
        return redirect('gestion:admin_panel')
    
    catalogo = CATALOGOS[tipo]
    item = get_object_or_404(catalogo['modelo'], pk=pk)
    
    item.activo = not item.activo
    item.save()
    
    estado = 'activado' if item.activo else 'desactivado'
    messages.success(request, f'{catalogo["nombre"]}: registro {estado}.')
    return redirect('gestion:admin_catalogo', tipo=tipo)

# ============================================================================
# PÁGINAS DE ERROR
# ============================================================================

def error_404(request, exception):
    return render(request, '404.html', status=404)

def error_403(request, exception):
    return render(request, '403.html', status=403)

def error_500(request):
    return render(request, '500.html', status=500)

@require_POST
@login_required
def verificar_password_fortaleza(request):
    """
    API para verificar fortaleza de contraseña en tiempo real.
    """
    password = request.POST.get('password', '')
    
    try:
        validate_password(password, user=request.user)
        return JsonResponse({
            'valida': True,
            'mensaje': 'Contraseña segura'
        })
    except ValidationError as e:
        return JsonResponse({
            'valida': False,
            'errores': e.messages
        })

# =============================================================================
# CONSENTIMIENTOS
# =============================================================================

@login_required
def consentimientos_lista(request):
    """Lista de todos los consentimientos"""
    consentimientos = Consentimiento.objects.select_related('persona', 'registrado_por').order_by('-fecha_registro')
    
    # Filtros
    persona_id = request.GET.get('persona', '')
    tipo = request.GET.get('tipo', '')
    estado = request.GET.get('estado', '')  # otorgado, revocado, pendiente
    
    if persona_id:
        consentimientos = consentimientos.filter(persona_id=persona_id)
    if tipo:
        consentimientos = consentimientos.filter(tipo=tipo)
    if estado:
        if estado == 'otorgado':
            consentimientos = consentimientos.filter(otorgado=True, fecha_revocacion__isnull=True)
        elif estado == 'revocado':
            consentimientos = consentimientos.filter(fecha_revocacion__isnull=False)
        elif estado == 'pendiente':
            consentimientos = consentimientos.filter(otorgado=False, fecha_revocacion__isnull=True)
    
    # Paginación
    paginator = Paginator(consentimientos, 20)
    page = request.GET.get('page', 1)
    try:
        consentimientos = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        consentimientos = paginator.page(1)
    
    context = {
        'consentimientos': consentimientos,
        'personas': Persona.objects.all().order_by('apellidos', 'nombres'),
        'tipos': Consentimiento.TIPO_CHOICES,
        'persona_filtro': persona_id,
        'tipo_filtro': tipo,
        'estado_filtro': estado,
    }
    return render(request, 'gestion/consentimientos_lista.html', context)


@login_required
def consentimiento_crear(request):
    """Crear nuevo consentimiento"""
    persona_id = request.GET.get('persona', '')
    
    if request.method == 'POST':
        # Crear consentimiento manualmente para evitar problemas con archivo vacío
        persona_id_post = request.POST.get('persona')
        tipo = request.POST.get('tipo')
        otorgado = request.POST.get('otorgado') == 'on'
        fecha_otorgamiento = request.POST.get('fecha_otorgamiento') or None
        observaciones = request.POST.get('observaciones', '')
        
        # Validaciones
        errores = []
        if not persona_id_post:
            errores.append('Debe seleccionar una persona.')
        if not tipo:
            errores.append('Debe seleccionar un tipo de consentimiento.')
        if otorgado and not fecha_otorgamiento:
            errores.append('Si el consentimiento fue otorgado, debe indicar la fecha.')
        
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            # Crear el consentimiento sin documento primero
            consentimiento = Consentimiento.objects.create(
                persona_id=persona_id_post,
                tipo=tipo,
                otorgado=otorgado,
                fecha_otorgamiento=fecha_otorgamiento,
                observaciones=observaciones,
                registrado_por=request.user
            )
            
            # Solo actualizar con archivo si se subió uno
            if request.FILES.get('documento_respaldo'):
                consentimiento.documento_respaldo = request.FILES['documento_respaldo']
                consentimiento.save(update_fields=['documento_respaldo'])
                messages.success(request, 'Consentimiento registrado exitosamente.')
            else:
                messages.success(request, 'Consentimiento registrado exitosamente (sin documento adjunto).')
            
            return redirect('gestion:consentimientos_lista')
    
    context = {
        'personas': Persona.objects.all().order_by('apellidos', 'nombres'),
        'tipos': Consentimiento.TIPO_CHOICES,
        'persona_preseleccionada': persona_id,
    }
    return render(request, 'gestion/consentimiento_form.html', context)


@login_required
def consentimiento_detalle(request, pk):
    """Ver detalle de un consentimiento"""
    consentimiento = get_object_or_404(Consentimiento, pk=pk)
    
    context = {
        'consentimiento': consentimiento,
    }
    return render(request, 'gestion/consentimiento_detalle.html', context)


@login_required
def consentimiento_editar(request, pk):
    """Editar un consentimiento"""
    consentimiento = get_object_or_404(Consentimiento, pk=pk)
    
    if request.method == 'POST':
        form = ConsentimientoForm(request.POST, request.FILES, instance=consentimiento)
        if form.is_valid():
            form.save()
            messages.success(request, 'Consentimiento actualizado exitosamente.')
            return redirect('gestion:consentimiento_detalle', pk=pk)
    else:
        form = ConsentimientoForm(instance=consentimiento)
    
    context = {
        'form': form,
        'consentimiento': consentimiento,
        'personas': Persona.objects.all().order_by('apellidos', 'nombres'),
        'tipos': Consentimiento.TIPO_CHOICES,
    }
    return render(request, 'gestion/consentimiento_form.html', context)


@login_required
def consentimiento_revocar(request, pk):
    """Revocar un consentimiento"""
    consentimiento = get_object_or_404(Consentimiento, pk=pk)
    
    if request.method == 'POST':
        consentimiento.fecha_revocacion = date.today()
        consentimiento.save()
        messages.success(request, 'Consentimiento revocado exitosamente.')
        return redirect('gestion:consentimiento_detalle', pk=pk)
    
    context = {
        'consentimiento': consentimiento,
    }
    return render(request, 'gestion/consentimiento_revocar.html', context)


@login_required
def persona_consentimientos(request, pk):
    """Ver consentimientos de una persona específica"""
    persona = get_object_or_404(Persona, pk=pk)
    consentimientos = persona.consentimientos.all().order_by('-fecha_registro')
    
    context = {
        'persona': persona,
        'consentimientos': consentimientos,
    }
    return render(request, 'gestion/persona_consentimientos.html', context)
